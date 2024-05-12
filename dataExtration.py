import openpyxl 
from openpyxl import Workbook


archivo = openpyxl.load_workbook('cedulas.xlsm')  #Cargar archivo
hoja = archivo['cedulas'] #Cargar hoja
nueva_hoja = archivo.create_sheet('ceds_p2')


columna_buscada = hoja['B'] #Columna donde se buscaran las unidades
unidades = ['Unidad: 555/AGUILA',
            'Unidad: 244/ALCATRAZ',
            'Unidad: 350/ANGELL',
            'Unidad: 485/ARCOIRIS II',
            'Unidad: 327/CAMILA',
            'Unidad: 221/CORAZON 1',
            'Unidad: 50/DESAFIO 1',
            'Unidad: 342/EMYORO',
            'Unidad: 262/EXITO',
            'Unidad: 344/FENIX 2',
            'Unidad: 310/GIRASOLES',
            'Unidad: 281/GRANDESA',
            'Unidad: 56/JUAN PABLO',
            'Unidad: 257/LUPITA 2',
            'Unidad: 349/LIBELULAS',
            'Unidad: 146/MAYA 1',
            'Unidad: 287/MERAKI',
            'Unidad: 197/NOVA',
            'Unidad: 328/NUEVO AMANECER',
            'Unidad: 502/ORQUIDEA',
            'Unidad: 168/PALOMA 2',
            'Unidad: 265/RENACIMIENTO',
            'Unidad: 237/RESPLANDOR II',
            'Unidad: 171/SATELITE 1', 
            'Unidad: 124/YARETZY',
            'Unidad: 556/ZOE',
            'Unidad: 411/28 DE ENERO',
            'Unidad: 353/GRATITUD']

nombre_limite = 'Total Unidad' #Identificador de que ahí terminan los datos de la unidad.


global inicio
global limite
global rango
datos_unidad = []
ini_cel = 0
columna_max = len(hoja[2])

def unidad(valor_c):
    if valor_c in unidades:
        return valor_c
    else:
        return False

nueva_hoja.append([])

for celda in columna_buscada:
    valor_celda = celda.value
    if valor_celda and valor_celda.startswith('Uni'):
        inicio = celda.row
        nombre = valor_celda
    elif celda.value == nombre_limite:
        limite = celda.row
        
        if unidad(nombre) !=  False:
            print(nombre)

            for row in hoja.iter_rows(min_row=inicio+2, max_row=limite-1, max_col=columna_max, values_only=True):
                updated_row = list(row)
                updated_row[0] = str(updated_row[0])
                updated_row[5] = nombre
                if updated_row[0] != 'None':
                    datos_unidad.append(updated_row) 
        
            #Ordenar datos_unidad en orden decreciente basándose en la primera columna
            datos_unidad.sort(key=lambda x: -int(x[0]) if x[0].isdigit() else float('inf'), reverse=True)

            for row in datos_unidad:
                if any(row):
                    nueva_hoja.append(row)

            calc_cel = ini_cel + 1
            nombre_cel = nueva_hoja['B' + str(calc_cel)]
            nombre_cel.value = nombre
            ini_cel = len(nueva_hoja['B'])
            datos_unidad.clear()
            nueva_hoja.append([])
            nueva_hoja.append([])

archivo.save('orden2.xlsx')